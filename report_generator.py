# report_generator.py - Generator Laporan Excel DISHUB DKI Jakarta
# ============================================================
# Menghasilkan laporan Excel multi-sheet yang siap cetak/kirim.
# Bisa dipanggil dari dashboard (tombol Export) atau CLI:
#
#   python report_generator.py                        # 30 hari, output otomatis
#   python report_generator.py --days 7               # 7 hari terakhir
#   python report_generator.py --output laporan.xlsx  # nama file custom
#
# Semua library GRATIS: openpyxl, pandas, sqlite3 (built-in)
# ============================================================

import sqlite3
import argparse
from datetime import datetime, timedelta
from io import BytesIO

import pandas as pd

# openpyxl is optional at lint time; provide a clear ImportError at runtime
try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.drawing.image import Image as XLImage
except ImportError as e:
    raise ImportError(
        "openpyxl is required to run report_generator.py. Install it with: pip install openpyxl"
    ) from e

try:
    from config import DB_PATH, CAMERA_LOCATIONS
except ImportError:
    DB_PATH = "violations.db"
    CAMERA_LOCATIONS = {
        "CAM_001": {"lat": -6.2088, "lon": 106.8456, "name": "Jl. Sudirman - Bundaran HI"},
        "CAM_002": {"lat": -6.1944, "lon": 106.8229, "name": "Jl. Thamrin - Monas"},
        "CAM_003": {"lat": -6.2297, "lon": 106.8295, "name": "Jl. Gatot Subroto"},
        "CAM_004": {"lat": -6.1750, "lon": 106.7972, "name": "Stasiun Grogol"},
        "CAM_005": {"lat": -6.2146, "lon": 106.8451, "name": "Pasar Tanah Abang"},
    }


# ============================================================
# KONSTANTA WARNA & STYLE
# ============================================================
# Warna resmi DISHUB / tema biru gelap profesional
C_NAVY       = "0F2A5C"   # Header utama
C_BLUE       = "1A56A0"   # Sub-header
C_LIGHT_BLUE = "D6E4F7"   # Baris alternating / zebra
C_YELLOW     = "F6C90E"   # Aksen / highlight
C_RED        = "C0392B"   # Pelanggaran tinggi / bahaya
C_ORANGE     = "E67E22"   # Pelanggaran sedang
C_GREEN      = "1E8449"   # Status aman / selesai
C_PURPLE     = "6C3483"   # Parkir liar
C_GREY_DARK  = "4A4A4A"
C_GREY_LIGHT = "F2F2F2"
C_WHITE      = "FFFFFF"

FONT_BASE    = "Arial"

VIOLATION_LABELS = {
    "busway_violation":       "Masuk Busway",
    "bicycle_lane_violation": "Masuk Jalur Sepeda",
    "illegal_parking":        "Parkir Liar",
}

# ============================================================
# HELPER STYLES
# ============================================================

def _border(style="thin"):
    s = Side(style=style, color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_font(size=11, bold=True, color=C_WHITE):
    return Font(name=FONT_BASE, size=size, bold=bold, color=color)

def _cell_font(size=10, bold=False, color=C_GREY_DARK):
    return Font(name=FONT_BASE, size=size, bold=bold, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def _apply_table_header(ws, row, cols, widths=None):
    """Tulis header baris dengan style navy."""
    for ci, col_name in enumerate(cols, 1):
        cell = ws.cell(row=row, column=ci, value=col_name)
        cell.font      = _hdr_font(10)
        cell.fill      = _fill(C_NAVY)
        cell.alignment = _align("center")
        cell.border    = _border()
        if widths:
            ws.column_dimensions[get_column_letter(ci)].width = widths[ci-1]

def _apply_data_row(ws, row, values, zebra=False, bold=False, color_map=None):
    """Tulis baris data dengan zebra striping opsional."""
    bg = C_LIGHT_BLUE if zebra else C_WHITE
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=ci, value=val)
        cell.font      = _cell_font(10, bold=bold)
        cell.fill      = _fill(bg)
        cell.border    = _border()
        cell.alignment = _align("left" if ci == 1 else "center")
        # Warna khusus kolom tertentu
        if color_map and ci in color_map:
            cell.font = _cell_font(10, bold=True, color=color_map[ci])

def _section_title(ws, row, col, text, span=1):
    """Judul section (bg biru muda, teks navy bold)."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = Font(name=FONT_BASE, size=11, bold=True, color=C_NAVY)
    cell.fill      = _fill(C_LIGHT_BLUE)
    cell.alignment = _align("left", wrap=False)
    cell.border    = _border()
    if span > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row,   end_column=col+span-1
        )


# ============================================================
# LOAD DATA FROM SQLITE
# ============================================================

def _load_data(db_path, days_back=30):
    """Ambil semua data yang diperlukan dari SQLite."""
    conn   = sqlite3.connect(db_path)
    cutoff = (datetime.now() - timedelta(days=days_back)).isoformat()

    df_violations = pd.read_sql_query("""
        SELECT id, timestamp, camera_id, camera_location,
               vehicle_type, license_plate, violation_type,
               zone_name, duration_seconds, status, etl_ticket_id
        FROM violations
        WHERE timestamp >= ?
        ORDER BY timestamp DESC
    """, conn, params=(cutoff,))

    df_vehicles = pd.read_sql_query("""
        SELECT license_plate, vehicle_type, first_seen, last_seen,
               total_violations, is_flagged
        FROM vehicles
        WHERE license_plate NOT IN ('', 'UNKNOWN')
        ORDER BY total_violations DESC
    """, conn)

    conn.close()
    return df_violations, df_vehicles


# ============================================================
# SHEET 1: COVER / RINGKASAN EKSEKUTIF
# ============================================================

def _sheet_cover(wb, df_v, df_k, days_back, generated_at):
    ws = wb.active
    ws.title = "📋 Ringkasan Eksekutif"
    ws.sheet_view.showGridLines = False

    # ---- Banner utama ----
    ws.merge_cells("A1:J3")
    banner = ws["A1"]
    banner.value     = "LAPORAN PELANGGARAN LALU LINTAS"
    banner.font      = Font(name=FONT_BASE, size=20, bold=True, color=C_WHITE)
    banner.fill      = _fill(C_NAVY)
    banner.alignment = _align("center", "center")
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20

    ws.merge_cells("A4:J4")
    sub = ws["A4"]
    sub.value     = "Dinas Perhubungan DKI Jakarta  ·  Traffic Enforcement System  ·  YOLOv8 + ByteTrack + EasyOCR"
    sub.font      = Font(name=FONT_BASE, size=10, italic=True, color=C_WHITE)
    sub.fill      = _fill(C_BLUE)
    sub.alignment = _align("center", "center")
    ws.row_dimensions[4].height = 18

    # ---- Meta info ----
    period_start = (datetime.now() - timedelta(days=days_back)).strftime("%d %B %Y")
    period_end   = datetime.now().strftime("%d %B %Y")
    meta = [
        ("Periode Laporan",   f"{period_start}  s/d  {period_end}"),
        ("Tanggal Dibuat",    generated_at.strftime("%d %B %Y, %H:%M WIB")),
        ("Dibuat Oleh",       "Sistem Otomatis — Traffic Enforcement System v1.1"),
        ("Jumlah Kamera",     str(len(CAMERA_LOCATIONS))),
        ("Status",            "RESMI / RAHASIA"),
    ]
    for i, (label, value) in enumerate(meta, 6):
        ws.merge_cells(f"A{i}:B{i}")
        lc = ws[f"A{i}"]
        lc.value     = label
        lc.font      = _hdr_font(10, color=C_NAVY)
        lc.fill      = _fill(C_LIGHT_BLUE)
        lc.alignment = _align("right")
        lc.border    = _border()

        ws.merge_cells(f"C{i}:F{i}")
        vc = ws[f"C{i}"]
        vc.value     = value
        vc.font      = _cell_font(10)
        vc.alignment = _align("left")
        vc.border    = _border()
        ws.row_dimensions[i].height = 16

    # ---- KPI Cards (baris 12-19) ----
    ws.merge_cells("A11:J11")
    kpi_hdr = ws["A11"]
    kpi_hdr.value     = "INDIKATOR KINERJA UTAMA"
    kpi_hdr.font      = Font(name=FONT_BASE, size=12, bold=True, color=C_WHITE)
    kpi_hdr.fill      = _fill(C_BLUE)
    kpi_hdr.alignment = _align("center")
    ws.row_dimensions[11].height = 18

    if df_v.empty:
        total, today_n, ticketed_n, unique_plates = 0, 0, 0, 0
        busway_n, bicycle_n, parking_n = 0, 0, 0
        avg_dur = 0.0
    else:
        total        = len(df_v)
        today_str    = datetime.now().strftime("%Y-%m-%d")
        today_n      = int((pd.to_datetime(df_v["timestamp"]).dt.strftime("%Y-%m-%d") == today_str).sum())
        ticketed_n   = int((df_v["status"] == "ticketed").sum())
        unique_plates = df_v["license_plate"].nunique()
        busway_n     = int((df_v["violation_type"] == "busway_violation").sum())
        bicycle_n    = int((df_v["violation_type"] == "bicycle_lane_violation").sum())
        parking_n    = int((df_v["violation_type"] == "illegal_parking").sum())
        park_df      = df_v[df_v["violation_type"] == "illegal_parking"]["duration_seconds"]
        avg_dur      = round(park_df.mean(), 1) if not park_df.empty else 0.0

    kpis = [
        ("Total Pelanggaran",        total,         C_YELLOW, "Semua jenis pelanggaran"),
        ("Sudah Ditilang (E-TLE)",   ticketed_n,    C_GREEN,  "Tiket berhasil diterbitkan"),
        ("Masuk Busway",             busway_n,      C_RED,    "Kendaraan masuk jalur bus"),
        ("Masuk Jalur Sepeda",       bicycle_n,     C_ORANGE, "Kendaraan masuk jalur sepeda"),
        ("Parkir Liar",              parking_n,     C_PURPLE, "Kendaraan parkir sembarangan"),
        ("Kendaraan Unik",           unique_plates, C_BLUE,   "Plat nomor berbeda"),
        ("Pelanggaran Hari Ini",     today_n,       C_NAVY,   "Deteksi hari ini"),
        (f"Rata-rata Durasi Parkir", f"{avg_dur}s", C_GREY_DARK, "Rata-rata durasi parkir liar"),
    ]

    for col_offset, (label, value, color, desc) in enumerate(kpis):
        col = col_offset + 1
        col_l = get_column_letter(col)
        # Baris label
        lc = ws.cell(row=12, column=col, value=label)
        lc.font      = Font(name=FONT_BASE, size=8, bold=True, color=C_WHITE)
        lc.fill      = _fill(color)
        lc.alignment = _align("center", "center", wrap=True)
        lc.border    = _border()
        ws.row_dimensions[12].height = 28

        # Baris nilai
        vc = ws.cell(row=13, column=col, value=value)
        vc.font      = Font(name=FONT_BASE, size=16, bold=True, color=color)
        vc.fill      = _fill(C_WHITE)
        vc.alignment = _align("center", "center")
        vc.border    = _border()
        ws.row_dimensions[13].height = 32

        # Baris deskripsi
        dc = ws.cell(row=14, column=col, value=desc)
        dc.font      = Font(name=FONT_BASE, size=7, color="888888")
        dc.fill      = _fill(C_GREY_LIGHT)
        dc.alignment = _align("center", "center", wrap=True)
        dc.border    = _border()
        ws.row_dimensions[14].height = 22

        ws.column_dimensions[col_l].width = 16

    # ---- Tren per tipe (ringkasan) ----
    ws.merge_cells("A16:J16")
    t2 = ws["A16"]
    t2.value     = "RINGKASAN PER JENIS PELANGGARAN"
    t2.font      = Font(name=FONT_BASE, size=11, bold=True, color=C_WHITE)
    t2.fill      = _fill(C_BLUE)
    t2.alignment = _align("center")
    ws.row_dimensions[16].height = 18

    summary_hdr = ["Jenis Pelanggaran", "Jumlah", "% dari Total",
                   "Sudah Ditilang", "Belum Ditilang", "Dominan Kendaraan"]
    _apply_table_header(ws, 17, summary_hdr, widths=[30,12,14,16,16,22])

    if not df_v.empty:
        for ri, vtype in enumerate(["busway_violation","bicycle_lane_violation","illegal_parking"]):
            sub = df_v[df_v["violation_type"] == vtype]
            n   = len(sub)
            pct = f"{n/total*100:.1f}%" if total > 0 else "0.0%"
            tkt = int((sub["status"] == "ticketed").sum())
            unt = n - tkt
            dom = sub["vehicle_type"].value_counts().idxmax() if not sub.empty else "-"
            _apply_data_row(ws, 18+ri,
                [VIOLATION_LABELS.get(vtype, vtype), n, pct, tkt, unt, dom],
                zebra=(ri % 2 == 1)
            )

    # Footer
    ws.merge_cells("A23:J23")
    ft = ws["A23"]
    ft.value     = f"Dokumen ini dibuat secara otomatis oleh Traffic Enforcement System DISHUB DKI Jakarta  |  {generated_at.strftime('%Y-%m-%d %H:%M:%S')}"
    ft.font      = Font(name=FONT_BASE, size=8, italic=True, color="999999")
    ft.alignment = _align("center")


# ============================================================
# SHEET 2: DATA PELANGGARAN LENGKAP
# ============================================================

def _sheet_violations(wb, df_v):
    ws = wb.create_sheet("📊 Data Pelanggaran")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    cols = ["No", "Waktu", "Kamera", "Lokasi", "Kendaraan",
            "Plat Nomor", "Jenis Pelanggaran", "Durasi (det)",
            "Status", "Nomor E-TLE"]
    widths = [6, 19, 10, 32, 12, 16, 24, 14, 12, 26]

    _apply_table_header(ws, 1, cols, widths=widths)

    STATUS_COLORS = {"ticketed": C_GREEN, "detected": C_RED}
    VTYPE_COLORS  = {
        "Masuk Busway":       C_RED,
        "Masuk Jalur Sepeda": C_ORANGE,
        "Parkir Liar":        C_PURPLE,
    }

    if df_v.empty:
        return

    for ri, (_, row) in enumerate(df_v.iterrows()):
        ts  = pd.to_datetime(row.get("timestamp","")).strftime("%d/%m/%Y %H:%M:%S") if row.get("timestamp") else "-"
        vl  = VIOLATION_LABELS.get(row.get("violation_type",""), row.get("violation_type",""))
        dur = row.get("duration_seconds", 0)
        dur_str = f"{int(dur//60)}m {int(dur%60)}s" if dur >= 60 else f"{dur:.0f}s"
        status  = row.get("status","detected")
        etl     = row.get("etl_ticket_id","") or "-"
        plate   = row.get("license_plate","UNKNOWN") or "UNKNOWN"

        values = [
            ri+1, ts,
            row.get("camera_id",""), row.get("camera_location",""),
            row.get("vehicle_type",""), plate,
            vl, dur_str, status.upper(), etl
        ]
        data_row = ri + 2
        zebra    = (ri % 2 == 1)
        bg       = C_LIGHT_BLUE if zebra else C_WHITE

        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=data_row, column=ci, value=val)
            cell.border    = _border()
            cell.alignment = _align("left" if ci in (4,7,10) else "center")

            # Default
            cell.font = _cell_font(9)
            cell.fill = _fill(bg)

            # Kolom status — warna berdasarkan nilai
            if ci == 9:
                color = STATUS_COLORS.get(status, C_GREY_DARK)
                cell.font = Font(name=FONT_BASE, size=9, bold=True, color=color)
            # Kolom jenis pelanggaran
            elif ci == 7:
                color = VTYPE_COLORS.get(vl, C_GREY_DARK)
                cell.font = Font(name=FONT_BASE, size=9, bold=False, color=color)
            # Plat UNKNOWN
            elif ci == 6 and plate == "UNKNOWN":
                cell.font = Font(name=FONT_BASE, size=9, italic=True, color="AAAAAA")

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(df_v)+1}"

    # Baris total
    total_row = len(df_v) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = _hdr_font(10, color=C_NAVY)
    ws.cell(row=total_row, column=1).fill = _fill(C_LIGHT_BLUE)
    ws.cell(row=total_row, column=2, value=f'=COUNTA(B2:B{total_row-1})').font = _hdr_font(10, color=C_NAVY)
    ws.cell(row=total_row, column=2).fill = _fill(C_LIGHT_BLUE)


# ============================================================
# SHEET 3: STATISTIK PER KAMERA & LOKASI
# ============================================================

def _sheet_cameras(wb, df_v):
    ws = wb.create_sheet("📷 Per Kamera")
    ws.sheet_view.showGridLines = False

    _section_title(ws, 1, 1, "STATISTIK PELANGGARAN PER KAMERA & LOKASI", span=8)
    ws.row_dimensions[1].height = 18

    cols   = ["Kamera ID", "Lokasi", "Total", "Busway",
              "Jalur Sepeda", "Parkir Liar", "Sudah Ditilang", "Jam Tersibuk"]
    widths = [12, 36, 10, 10, 14, 12, 16, 14]
    _apply_table_header(ws, 2, cols, widths=widths)

    if df_v.empty:
        return

    cam_stats = []
    for cam_id, info in CAMERA_LOCATIONS.items():
        sub     = df_v[df_v["camera_id"] == cam_id]
        if sub.empty:
            continue
        total   = len(sub)
        busway  = int((sub["violation_type"] == "busway_violation").sum())
        bicycle = int((sub["violation_type"] == "bicycle_lane_violation").sum())
        parking = int((sub["violation_type"] == "illegal_parking").sum())
        tkt     = int((sub["status"] == "ticketed").sum())
        hour_s  = pd.to_datetime(sub["timestamp"]).dt.hour.value_counts()
        busiest = f"{hour_s.idxmax():02d}:00" if not hour_s.empty else "-"
        cam_stats.append((cam_id, info["name"], total, busway, bicycle, parking, tkt, busiest))

    cam_stats.sort(key=lambda x: x[2], reverse=True)

    for ri, row_vals in enumerate(cam_stats):
        _apply_data_row(ws, ri+3, list(row_vals), zebra=(ri % 2 == 1))

    # Totals row
    total_data_row = len(cam_stats) + 3
    ws.cell(total_data_row, 1, "TOTAL").font = _hdr_font(10, color=C_NAVY)
    ws.cell(total_data_row, 1).fill = _fill(C_LIGHT_BLUE)
    for col_i in range(3, 8):
        col_l = get_column_letter(col_i)
        cell  = ws.cell(total_data_row, col_i,
                        value=f"=SUM({col_l}3:{col_l}{total_data_row-1})")
        cell.font = _hdr_font(10, color=C_NAVY)
        cell.fill = _fill(C_LIGHT_BLUE)
        cell.border = _border()

    # ---- Chart Batang Per Kamera ----
    chart = BarChart()
    chart.type    = "col"
    chart.grouping = "clustered"
    chart.title   = "Pelanggaran per Kamera"
    chart.y_axis.title = "Jumlah"
    chart.x_axis.title = "Kamera"
    chart.style   = 10
    chart.width   = 22
    chart.height  = 14

    # Data: kolom Total, Busway, Jalur Sepeda, Parkir Liar (col 3-6)
    data  = Reference(ws, min_col=3, max_col=6,
                      min_row=2, max_row=2+len(cam_stats))
    cats  = Reference(ws, min_col=2, min_row=3, max_row=2+len(cam_stats))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    series_colors = [C_BLUE, C_RED, C_ORANGE, C_PURPLE]
    for i, clr in enumerate(series_colors):
        if i < len(chart.series):
            chart.series[i].graphicalProperties.solidFill = clr

    ws.add_chart(chart, f"A{total_data_row+2}")


# ============================================================
# SHEET 4: ANALISIS JAM & HARI
# ============================================================

def _sheet_time_analysis(wb, df_v):
    ws = wb.create_sheet("⏰ Pola Waktu")
    ws.sheet_view.showGridLines = False

    # ---- Tabel per Jam ----
    _section_title(ws, 1, 1, "DISTRIBUSI PELANGGARAN PER JAM", span=5)
    cols_h = ["Jam", "Total", "Busway", "Jalur Sepeda", "Parkir Liar"]
    widths_h = [10, 10, 12, 14, 12]
    _apply_table_header(ws, 2, cols_h, widths=widths_h)

    if not df_v.empty:
        df_v2 = df_v.copy()
        df_v2["hour"] = pd.to_datetime(df_v2["timestamp"]).dt.hour
        for ri, hour in enumerate(range(24)):
            sub  = df_v2[df_v2["hour"] == hour]
            vals = [
                f"{hour:02d}:00",
                len(sub),
                int((sub["violation_type"] == "busway_violation").sum()),
                int((sub["violation_type"] == "bicycle_lane_violation").sum()),
                int((sub["violation_type"] == "illegal_parking").sum()),
            ]
            _apply_data_row(ws, ri+3, vals, zebra=(ri % 2 == 1))

        total_row_h = 27
        ws.cell(total_row_h, 1, "TOTAL").font = _hdr_font(10, color=C_NAVY)
        ws.cell(total_row_h, 1).fill = _fill(C_LIGHT_BLUE)
        for ci in range(2, 6):
            cl = get_column_letter(ci)
            c  = ws.cell(total_row_h, ci, value=f"=SUM({cl}3:{cl}26)")
            c.font  = _hdr_font(10, color=C_NAVY)
            c.fill  = _fill(C_LIGHT_BLUE)
            c.border = _border()

    # ---- Tabel per Hari (kolom G ke kanan) ----
    _section_title(ws, 1, 7, "DISTRIBUSI PER HARI DALAM SEMINGGU", span=4)
    day_order = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
    day_id    = ["Senin","Selasa","Rabu","Kamis","Jumat","Sabtu","Minggu"]
    cols_d    = ["Hari", "Total", "% dari Minggu", "Dominan Jenis"]
    widths_d  = [12, 10, 14, 24]
    _apply_table_header(ws, 2, cols_d,
                        widths=[widths_d[i] for i in range(4)])
    # Re-apply for columns G+
    for ci, (col_name, w) in enumerate(zip(cols_d, widths_d), 7):
        cell = ws.cell(row=2, column=ci, value=col_name)
        cell.font      = _hdr_font(10)
        cell.fill      = _fill(C_NAVY)
        cell.alignment = _align("center")
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w

    if not df_v.empty:
        df_v3 = df_v.copy()
        df_v3["ts"]  = pd.to_datetime(df_v3["timestamp"])
        df_v3["day"] = df_v3["ts"].dt.day_name()
        grand_total  = len(df_v3)
        for ri, (eng, ind) in enumerate(zip(day_order, day_id)):
            sub  = df_v3[df_v3["day"] == eng]
            n    = len(sub)
            pct  = f"{n/grand_total*100:.1f}%" if grand_total > 0 else "0.0%"
            dom  = VIOLATION_LABELS.get(
                sub["violation_type"].value_counts().idxmax()
                if not sub.empty else "", "-"
            )
            row_vals = [ind, n, pct, dom]
            zebra = (ri % 2 == 1)
            bg    = C_LIGHT_BLUE if zebra else C_WHITE
            for ci, val in enumerate(row_vals, 7):
                cell = ws.cell(ri+3, ci, value=val)
                cell.font      = _cell_font(10)
                cell.fill      = _fill(bg)
                cell.border    = _border()
                cell.alignment = _align("center" if ci > 7 else "left")

    # ---- Chart Jam ----
    chart = LineChart()
    chart.title   = "Tren Pelanggaran per Jam"
    chart.y_axis.title = "Jumlah"
    chart.x_axis.title = "Jam"
    chart.style   = 10
    chart.width   = 22
    chart.height  = 14
    data_ref  = Reference(ws, min_col=2, max_col=5, min_row=2, max_row=26)
    cats_ref  = Reference(ws, min_col=1, min_row=3, max_row=26)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "A29")


# ============================================================
# SHEET 5: REKAP HARIAN (TREN)
# ============================================================

def _sheet_daily_trend(wb, df_v):
    ws = wb.create_sheet("📈 Tren Harian")
    ws.sheet_view.showGridLines = False

    _section_title(ws, 1, 1, "TREN PELANGGARAN HARIAN", span=7)

    cols   = ["Tanggal", "Total", "Busway", "Jalur Sepeda",
              "Parkir Liar", "Sudah Ditilang", "Kumulatif"]
    widths = [14, 10, 12, 14, 12, 16, 14]
    _apply_table_header(ws, 2, cols, widths=widths)

    if df_v.empty:
        return

    df_v2 = df_v.copy()
    df_v2["date"] = pd.to_datetime(df_v2["timestamp"]).dt.date
    daily = df_v2.groupby("date").apply(lambda g: pd.Series({
        "total":   len(g),
        "busway":  int((g["violation_type"]=="busway_violation").sum()),
        "bicycle": int((g["violation_type"]=="bicycle_lane_violation").sum()),
        "parking": int((g["violation_type"]=="illegal_parking").sum()),
        "tkt":     int((g["status"]=="ticketed").sum()),
    })).reset_index().sort_values("date")

    daily["cumul"] = daily["total"].cumsum()

    for ri, (_, row) in enumerate(daily.iterrows()):
        _apply_data_row(ws, ri+3, [
            str(row["date"]),
            int(row["total"]), int(row["busway"]),
            int(row["bicycle"]), int(row["parking"]),
            int(row["tkt"]),    int(row["cumul"]),
        ], zebra=(ri % 2 == 1))

    # Totals
    n_data = len(daily)
    tr     = n_data + 3
    ws.cell(tr, 1, "TOTAL").font  = _hdr_font(10, color=C_NAVY)
    ws.cell(tr, 1).fill           = _fill(C_LIGHT_BLUE)
    for ci in range(2, 7):
        cl = get_column_letter(ci)
        c  = ws.cell(tr, ci, value=f"=SUM({cl}3:{cl}{tr-1})")
        c.font  = _hdr_font(10, color=C_NAVY)
        c.fill  = _fill(C_LIGHT_BLUE)
        c.border = _border()

    # Line chart
    chart = LineChart()
    chart.title   = "Tren Harian Pelanggaran"
    chart.y_axis.title = "Jumlah"
    chart.style   = 10
    chart.width   = 26
    chart.height  = 14
    data_ref = Reference(ws, min_col=2, max_col=5,
                         min_row=2, max_row=2+n_data)
    cats_ref = Reference(ws, min_col=1, min_row=3, max_row=2+n_data)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, f"A{tr+2}")


# ============================================================
# SHEET 6: RECIDIVISM — PELANGGAR BERULANG
# ============================================================

def _sheet_recidivism(wb, df_k):
    ws = wb.create_sheet("🚨 Pelanggar Berulang")
    ws.sheet_view.showGridLines = False

    _section_title(ws, 1, 1, "DAFTAR KENDARAAN PELANGGAR BERULANG (RECIDIVISM)", span=7)

    cols   = ["Plat Nomor", "Jenis Kendaraan", "Total Pelanggaran",
              "Level Risiko", "Pertama Terlihat", "Terakhir Terlihat", "Ditandai"]
    widths = [18, 16, 18, 16, 22, 22, 12]
    _apply_table_header(ws, 2, cols, widths=widths)

    if df_k.empty:
        return

    RISK_COLORS = {"HIGH": C_RED, "MEDIUM": C_ORANGE, "LOW": C_GREEN}

    for ri, (_, row) in enumerate(df_k.iterrows()):
        n     = int(row.get("total_violations", 0))
        risk  = "HIGH" if n >= 5 else "MEDIUM" if n >= 3 else "LOW"
        emoji = "🔴" if risk=="HIGH" else "🟠" if risk=="MEDIUM" else "🟢"
        flagged = "⚑ DITANDAI" if int(row.get("is_flagged", 0)) else "-"

        data_row = ri + 3
        zebra    = (ri % 2 == 1)
        bg       = C_LIGHT_BLUE if zebra else C_WHITE
        values   = [
            row.get("license_plate",""),
            row.get("vehicle_type",""),
            n,
            f"{emoji} {risk}",
            str(row.get("first_seen",""))[:19],
            str(row.get("last_seen",""))[:19],
            flagged,
        ]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(data_row, ci, value=val)
            cell.fill   = _fill(bg)
            cell.border = _border()
            cell.alignment = _align("center" if ci != 1 else "left")

            if ci == 3:  # total violations
                clr = RISK_COLORS.get(risk, C_GREY_DARK)
                cell.font = Font(name=FONT_BASE, size=10, bold=True, color=clr)
            elif ci == 4:
                clr = RISK_COLORS.get(risk, C_GREY_DARK)
                cell.font = Font(name=FONT_BASE, size=10, bold=True, color=clr)
            elif ci == 7 and flagged != "-":
                cell.font = Font(name=FONT_BASE, size=10, bold=True, color=C_RED)
            else:
                cell.font = _cell_font(10)

    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{len(df_k)+2}"

    # Pie chart: distribusi risiko
    risk_counts = {"HIGH":0, "MEDIUM":0, "LOW":0}
    for _, row in df_k.iterrows():
        n    = int(row.get("total_violations",0))
        risk = "HIGH" if n >= 5 else "MEDIUM" if n >= 3 else "LOW"
        risk_counts[risk] += 1

    # Tulis data bantu untuk pie
    pie_row = len(df_k) + 5
    ws.cell(pie_row,   1, "Level").font = _hdr_font(10)
    ws.cell(pie_row,   2, "Jumlah").font = _hdr_font(10)
    ws.cell(pie_row+1, 1, "HIGH").font   = _cell_font(10)
    ws.cell(pie_row+1, 2, risk_counts["HIGH"])
    ws.cell(pie_row+2, 1, "MEDIUM").font = _cell_font(10)
    ws.cell(pie_row+2, 2, risk_counts["MEDIUM"])
    ws.cell(pie_row+3, 1, "LOW").font    = _cell_font(10)
    ws.cell(pie_row+3, 2, risk_counts["LOW"])

    pie = PieChart()
    pie.title  = "Distribusi Level Risiko"
    pie.style  = 10
    pie.width  = 16
    pie.height = 12
    pie_data = Reference(ws, min_col=2, min_row=pie_row,   max_row=pie_row+3)
    pie_cats = Reference(ws, min_col=1, min_row=pie_row+1, max_row=pie_row+3)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)
    slices = [DataPoint(idx=0), DataPoint(idx=1), DataPoint(idx=2)]
    slices[0].graphicalProperties.solidFill = C_RED
    slices[1].graphicalProperties.solidFill = C_ORANGE
    slices[2].graphicalProperties.solidFill = C_GREEN
    pie.series[0].dPt = slices
    ws.add_chart(pie, f"D{pie_row}")


# ============================================================
# SHEET 7: STATUS E-TLE
# ============================================================

def _sheet_etle(wb, df_v):
    ws = wb.create_sheet("🎫 Status E-TLE")
    ws.sheet_view.showGridLines = False

    _section_title(ws, 1, 1, "LAPORAN STATUS ELECTRONIC TRAFFIC LAW ENFORCEMENT (E-TLE)", span=8)

    cols   = ["No", "Nomor E-TLE", "Waktu Pelanggaran", "Lokasi",
              "Plat Nomor", "Kendaraan", "Jenis Pelanggaran", "Status"]
    widths = [6, 28, 20, 32, 16, 12, 24, 14]
    _apply_table_header(ws, 2, cols, widths=widths)

    if df_v.empty:
        return

    df_tkt = df_v[df_v["status"] == "ticketed"].copy()
    df_det = df_v[df_v["status"] == "detected"].copy()
    df_ordered = pd.concat([df_tkt, df_det]).reset_index(drop=True)

    for ri, (_, row) in enumerate(df_ordered.iterrows()):
        ts     = pd.to_datetime(row.get("timestamp","")).strftime("%d/%m/%Y %H:%M:%S") if row.get("timestamp") else "-"
        status = row.get("status","detected")
        etl    = row.get("etl_ticket_id","") or "— Belum Diterbitkan —"
        vl     = VIOLATION_LABELS.get(row.get("violation_type",""), row.get("violation_type",""))

        data_row = ri + 3
        zebra    = (ri % 2 == 1)
        bg       = C_LIGHT_BLUE if zebra else C_WHITE

        values = [
            ri+1, etl, ts,
            row.get("camera_location",""),
            row.get("license_plate","UNKNOWN"),
            row.get("vehicle_type",""),
            vl, "DITILANG ✓" if status=="ticketed" else "BELUM DITILANG"
        ]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(data_row, ci, value=val)
            cell.fill      = _fill(bg)
            cell.border    = _border()
            cell.alignment = _align("center" if ci != 4 else "left")

            if ci == 8:
                if status == "ticketed":
                    cell.font = Font(name=FONT_BASE, size=10, bold=True, color=C_GREEN)
                else:
                    cell.font = Font(name=FONT_BASE, size=10, color=C_RED)
            elif ci == 2 and etl != "— Belum Diterbitkan —":
                cell.font = Font(name=FONT_BASE, size=9, bold=True, color=C_BLUE)
            else:
                cell.font = _cell_font(9)

    # Summary bawah
    total_v = len(df_ordered)
    sum_row  = total_v + 4
    ws.merge_cells(f"A{sum_row}:B{sum_row}")
    ws.cell(sum_row, 1, "Sudah Ditilang").font  = _hdr_font(10, color=C_GREEN)
    ws.cell(sum_row, 1).fill = _fill(C_LIGHT_BLUE)
    ws.cell(sum_row, 3, f'=COUNTIF(H3:H{sum_row-1},"DITILANG ✓")').font = _hdr_font(10, color=C_GREEN)
    ws.cell(sum_row, 3).fill = _fill(C_LIGHT_BLUE)
    ws.cell(sum_row, 3).border = _border()

    ws.merge_cells(f"A{sum_row+1}:B{sum_row+1}")
    ws.cell(sum_row+1, 1, "Belum Ditilang").font = _hdr_font(10, color=C_RED)
    ws.cell(sum_row+1, 1).fill = _fill(C_LIGHT_BLUE)
    ws.cell(sum_row+1, 3, f'=COUNTIF(H3:H{sum_row-1},"BELUM DITILANG")').font = _hdr_font(10, color=C_RED)
    ws.cell(sum_row+1, 3).fill = _fill(C_LIGHT_BLUE)
    ws.cell(sum_row+1, 3).border = _border()

    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{total_v+2}"


# ============================================================
# MAIN FUNCTION
# ============================================================

def generate_report(
    db_path: str = None,
    output_path: str = None,
    days_back: int = 30,
    to_bytes: bool = False,
) -> str | bytes:
    """
    Generate laporan Excel lengkap.

    Args:
        db_path     : Path ke violations.db. Default: DB_PATH dari config.
        output_path : Path file output .xlsx. Default: Laporan_DISHUB_YYYYMMDD.xlsx
        days_back   : Ambil data N hari ke belakang.
        to_bytes    : True → kembalikan bytes (untuk st.download_button di Streamlit),
                      False → simpan ke file dan kembalikan path.

    Returns:
        Path file jika to_bytes=False, bytes jika to_bytes=True.
    """
    if db_path is None:
        db_path = DB_PATH
    if output_path is None:
        output_path = f"Laporan_DISHUB_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    generated_at = datetime.now()
    print(f"[*] Loading data dari {db_path} ({days_back} hari terakhir)...")
    df_violations, df_vehicles = _load_data(db_path, days_back)
    print(f"[✓] {len(df_violations)} pelanggaran, {len(df_vehicles)} kendaraan unik.")

    wb = Workbook()

    print("[*] Membuat sheet: Ringkasan Eksekutif...")
    _sheet_cover(wb, df_violations, df_vehicles, days_back, generated_at)

    print("[*] Membuat sheet: Data Pelanggaran...")
    _sheet_violations(wb, df_violations)

    print("[*] Membuat sheet: Per Kamera...")
    _sheet_cameras(wb, df_violations)

    print("[*] Membuat sheet: Pola Waktu...")
    _sheet_time_analysis(wb, df_violations)

    print("[*] Membuat sheet: Tren Harian...")
    _sheet_daily_trend(wb, df_violations)

    print("[*] Membuat sheet: Pelanggar Berulang...")
    _sheet_recidivism(wb, df_vehicles)

    print("[*] Membuat sheet: Status E-TLE...")
    _sheet_etle(wb, df_violations)

    if to_bytes:
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        print("[✓] Laporan selesai (bytes mode).")
        return buf.read()
    else:
        wb.save(output_path)
        print(f"[✓] Laporan disimpan ke: {output_path}")
        return output_path


# ============================================================
# CLI
# ============================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generator Laporan Excel — DISHUB DKI Jakarta Traffic Enforcement System"
    )
    parser.add_argument("--days",   type=int, default=30,
                        help="Rentang data (hari ke belakang, default: 30)")
    parser.add_argument("--db",     default=None,
                        help=f"Path database SQLite (default: {DB_PATH})")
    parser.add_argument("--output", default=None,
                        help="Nama file output .xlsx (default: Laporan_DISHUB_YYYYMMDD.xlsx)")
    args = parser.parse_args()

    path = generate_report(
        db_path=args.db,
        output_path=args.output,
        days_back=args.days,
    )
    print(f"\n[→] Buka file: {path}")
    print("[→] Atau jalankan dashboard: streamlit run dashboard.py")
